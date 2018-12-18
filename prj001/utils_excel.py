
from openpyxl import Workbook


def create_xls(s):
    """
    创建xls文件
    :param s: 序列化器对象
    :return: 文件地址
    """
    wb = Workbook()
    # 先抽离出五个表
    r = 1
    c = 1
    for i, obj in enumerate(s.data):
        ws = wb.create_sheet(title=obj.get('name'), index=i)
        for k, v in obj.items():

            if not isinstance(v, type(None)):

                if isinstance(obj[k], str) or isinstance(obj[k], bool) or isinstance(obj[k], int):
                    # 正常
                    ws.cell(row=r, column=c, value=k)
                    ws.cell(row=r, column=c + 1, value=v)
                    r += 1
                else:
                    # 月经
                    for index, value in obj[k].items():
                        ws.cell(row=r, column=c, value=index)
                        ws.cell(row=r, column=c+1, value=value)
                        r += 1
            else:
                ws.cell(row=r, column=c, value='该患者无此相关数据')
                r += 1
        r, n = 1, 1
    wb.save("./a.xlsx")
    return

